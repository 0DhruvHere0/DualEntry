from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
def create_financial_report_excel(
    trial_balance,
    ledgers,
    income_statement,
    balance_sheet,
):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    trial_sheet = workbook.create_sheet("Trial Balance")
    trial_sheet.append([
        "Account ID",
        "Account Name",
        "Category",
        "Debit",
        "Credit",
    ])
    for cell in trial_sheet[1]:
        cell.font = Font(bold=True)
    for account in trial_balance.accounts:
        trial_sheet.append([
            account.account_id,
            account.account_name,
            account.category,
            account.debit,
            account.credit,
        ])
    trial_sheet.append([
        "",
        "TOTAL",
        "",
        trial_balance.total_debit,
        trial_balance.total_credit,
    ])
    for cell in trial_sheet[trial_sheet.max_row]:
        cell.font = Font(bold=True)
    ledger_sheet = workbook.create_sheet("Account Ledgers")
    for ledger in ledgers:
        ledger_sheet.append([
            f"Account: {ledger.account_name}"
        ])
        for cell in ledger_sheet[ledger_sheet.max_row]:
            cell.font = Font(
                bold=True,
                size=14,
            )
        ledger_sheet.append([
            "Entry ID",
            "Transaction ID",
            "Date",
            "Description",
            "Transaction Type",
            "Entry Type",
            "Debit",
            "Credit",
            "Balance",
        ])
        for cell in ledger_sheet[ledger_sheet.max_row]:
            cell.font = Font(bold=True)
        for entry in ledger.entries:
            ledger_sheet.append([
                entry.entry_id,
                entry.transaction_id,
                entry.created_at,
                entry.description,
                entry.transaction_type,
                entry.entry_type,
                entry.debit,
                entry.credit,
                entry.balance,
            ])
        ledger_sheet.append([])
    income_sheet = workbook.create_sheet("Income Statement")
    income_sheet.append(["INCOME"])
    for cell in income_sheet[1]:
        cell.font = Font(bold=True, size=14)
    income_sheet.append([
        "Account ID",
        "Account Name",
        "Amount",
    ])
    for cell in income_sheet[2]:
        cell.font = Font(bold=True)
    for item in income_statement.income:
        income_sheet.append([
            item.account_id,
            item.account_name,
            item.amount,
        ])
    income_sheet.append([
        "",
        "TOTAL INCOME",
        income_statement.total_income,
    ])
    for cell in income_sheet[income_sheet.max_row]:
        cell.font = Font(bold=True)
    income_sheet.append([])
    income_sheet.append(["EXPENSES"])
    for cell in income_sheet[income_sheet.max_row]:
        cell.font = Font(
            bold=True,
            size=14,
        )
    income_sheet.append([
        "Account ID",
        "Account Name",
        "Amount",
    ])
    for cell in income_sheet[income_sheet.max_row]:
        cell.font = Font(bold=True)
    for item in income_statement.expenses:
        income_sheet.append([
            item.account_id,
            item.account_name,
            item.amount,
        ])
    income_sheet.append([
        "",
        "TOTAL EXPENSES",
        income_statement.total_expenses,
    ])
    for cell in income_sheet[income_sheet.max_row]:
        cell.font = Font(bold=True)
    income_sheet.append([
        "",
        "NET INCOME",
        income_statement.net_income,
    ])
    for cell in income_sheet[income_sheet.max_row]:
        cell.font = Font(bold=True)
    balance_sheet_page = workbook.create_sheet("Balance Sheet")
    balance_sheet_page.append(["ASSETS"])
    for cell in balance_sheet_page[1]:
        cell.font = Font(bold=True, size=14)
    balance_sheet_page.append([
        "Account ID",
        "Account Name",
        "Amount",
    ])
    for cell in balance_sheet_page[2]:
        cell.font = Font(bold=True)
    for item in balance_sheet.assets:
        balance_sheet_page.append([
            item.account_id,
            item.account_name,
            item.amount,
        ])
    balance_sheet_page.append([
        "",
        "TOTAL ASSETS",
        balance_sheet.total_assets,
    ])
    for cell in balance_sheet_page[
        balance_sheet_page.max_row
    ]:
        cell.font = Font(bold=True)
    balance_sheet_page.append([])
    balance_sheet_page.append(["LIABILITIES"])
    for cell in balance_sheet_page[
        balance_sheet_page.max_row
    ]:
        cell.font = Font(
            bold=True,
            size=14,
    )
    balance_sheet_page.append([
        "Account ID",
        "Account Name",
        "Amount",
    ])
    for cell in balance_sheet_page[
        balance_sheet_page.max_row
    ]:
        cell.font = Font(bold=True)
    for item in balance_sheet.liabilities:
        balance_sheet_page.append([
            item.account_id,
            item.account_name,
            item.amount,
        ])
    balance_sheet_page.append([
        "",
        "TOTAL LIABILITIES",
        balance_sheet.total_liabilities,
    ])
    for cell in balance_sheet_page[
        balance_sheet_page.max_row
    ]:
        cell.font = Font(bold=True)
    balance_sheet_page.append([])
    balance_sheet_page.append(["EQUITY"])
    for cell in balance_sheet_page[
        balance_sheet_page.max_row
    ]:
        cell.font = Font(
            bold=True,
            size=14,
    )
    balance_sheet_page.append([
        "Account ID",
        "Account Name",
        "Amount",
    ])
    for cell in balance_sheet_page[
        balance_sheet_page.max_row
    ]:
        cell.font = Font(bold=True)
    for item in balance_sheet.equity:
        balance_sheet_page.append([
            item.account_id,
            item.account_name,
            item.amount,
        ])
    balance_sheet_page.append([
        "",
        "TOTAL EQUITY",
        balance_sheet.total_equity,
    ])
    for cell in balance_sheet_page[
        balance_sheet_page.max_row
    ]:
        cell.font = Font(bold=True)
    balance_sheet_page.append([
        "",
        "TOTAL LIABILITIES + EQUITY",
        balance_sheet.total_liabilities_and_equity,
    ])
    for cell in balance_sheet_page[
        balance_sheet_page.max_row
    ]:
        cell.font = Font(bold=True)
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            for cell in column:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )
            column_letter = get_column_letter(
                column[0].column
            )
            sheet.column_dimensions[
                column_letter
            ].width = min(max_length + 2, 40)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center"
                )
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output